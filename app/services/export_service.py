import csv
import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.work_log import WorkLog


class ExportService:
    """Service for exporting work logs to CSV and Excel."""

    COLUMNS = [
        ('Date', 'date'),
        ('Time', 'created_at'),
        ('Annotator', 'annotator_name'),
        ('Project', 'project_name'),
        ('Task Type', 'task_type'),
        ('Images Done', 'images_done'),
        ('Hours Spent', 'hours_spent'),
        ('Status', 'status'),
        ('Challenges', 'challenges'),
        ('Suggestions', 'suggestions'),
        ('Extra Notes', 'extra_notes'),
    ]

    @classmethod
    def generate_csv(cls, logs: List[WorkLog]) -> str:
        """Generate CSV string from work logs."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([col[0] for col in cls.COLUMNS])
        
        # Write data
        for log in logs:
            row = []
            for _, attr in cls.COLUMNS:
                value = getattr(log, attr, '')
                if attr == 'created_at':
                    # Format time only
                    value = value[11:16] if len(value) > 16 else value
                row.append(value)
            writer.writerow(row)
        
        return output.getvalue()

    @classmethod
    def generate_excel(cls, logs: List[WorkLog], date: str) -> bytes:
        """Generate Excel file from work logs."""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Report {date}"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        # Write header
        for col_idx, (col_name, _) in enumerate(cls.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row_idx, log in enumerate(logs, 2):
            for col_idx, (_, attr) in enumerate(cls.COLUMNS, 1):
                value = getattr(log, attr, '')
                if attr == 'created_at':
                    value = value[11:16] if len(value) > 16 else value
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                
                # Right-align numeric columns
                if attr in ('images_done', 'hours_spent'):
                    cell.alignment = Alignment(horizontal="right")
        
        # Auto-adjust column widths
        for col_idx, (col_name, attr) in enumerate(cls.COLUMNS, 1):
            max_length = len(col_name)
            for row_idx in range(2, len(logs) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Cap at reasonable width
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

